from django.shortcuts import render
from openpyxl import load_workbook
from datetime import datetime

def index(request):
    return render(request, 'automatizacoes/index.html')

def executar(request):
    caminho_completo = './colaboradores.xlsx'
    colaboradores = load_workbook(caminho_completo)
    print(colaboradores.sheetnames)

    aba_inicial = colaboradores['Worksheet']

    ultima_linha = aba_inicial.max_row

    for linha in range(1, ultima_linha + 1):
        data_formatada = datetime.strptime(aba_inicial.cell(row=linha, column=3).value, '%Y-%m-%d %H:%M:%S')
        aba = nome_da_aba(data_formatada.strftime('%m'))

        if aba not in colaboradores.sheetnames:
            colaboradores.create_sheet(aba)

        aba_destino = colaboradores[aba]
        max_linha = aba_destino.max_row + 1
        aba_destino.cell(row=max_linha, column=1).value = aba_inicial.cell(row=linha, column=1).value
        aba_destino.cell(row=max_linha, column=2).value = aba_inicial.cell(row=linha, column=2).value
        aba_destino.cell(row=max_linha, column=3).value = aba_inicial.cell(row=linha, column=3).value

    colaboradores.save("./colaboradores_separados.xlsx")

    return render(request, 'automatizacoes/sucesso.html')

def nome_da_aba(mes):
    meses = {
        '01': 'Janeiro',
        '02': 'Fevereiro',
        '03': 'Março',
        '04': 'Abril',
        '05': 'Maio',
        '06': 'Junho',
        '07': 'Julho',
        '08': 'Agosto',
        '09': 'Setembro',
        '10': 'Outubro',
        '11': 'Novembro',
        '12': 'Dezembro'
    }
    return meses[mes]